import os
import csv
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import platform
import datetime

class FoliarioApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Generatore di Foliari Pro")
        self.root.geometry("1000x650")
        
        # Dizionario per tenere traccia dello stato
        self.included_items = {}
        
        self.setup_ui()

    def setup_ui(self):
        # --- PANNELLO LATERALE DESTRO (Comandi e Statistiche) ---
        right_frame = tk.Frame(self.root, width=250, pady=10, padx=10, bg="#f0f0f0")
        right_frame.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Titolo comandi
        tk.Label(right_frame, text="COMANDI", font=("Helvetica", 12, "bold"), bg="#f0f0f0").pack(pady=(0, 10))
        
        # Bottoni principali
        tk.Button(right_frame, text="📁 1. Seleziona Cartella", command=self.load_directory, height=2).pack(fill=tk.X, pady=5)
        
        tk.Label(right_frame, text="Esportazione:", bg="#f0f0f0").pack(pady=(15, 0), anchor="w")
        tk.Button(right_frame, text="📄 2. Esporta CSV", command=self.export_csv).pack(fill=tk.X, pady=2)
        tk.Button(right_frame, text="📊 3. Esporta Excel (.xlsx)", command=self.export_excel).pack(fill=tk.X, pady=2)
        
        # Separatore
        ttk.Separator(right_frame, orient='horizontal').pack(fill=tk.X, pady=15)
        
        # Controlli visualizzazione
        tk.Label(right_frame, text="Visualizzazione:", bg="#f0f0f0").pack(anchor="w")
        tk.Button(right_frame, text="➕ Espandi Tutto", command=self.expand_all).pack(fill=tk.X, pady=2)
        tk.Button(right_frame, text="➖ Comprimi Tutto", command=self.collapse_all).pack(fill=tk.X, pady=2)
        
        # Istruzioni
        tk.Label(right_frame, text="💡 Doppio clic su una riga\nper Includere/Escludere", fg="#555555", bg="#f0f0f0", justify=tk.LEFT).pack(pady=15)
        
        # Separatore e Statistiche
        ttk.Separator(right_frame, orient='horizontal').pack(fill=tk.X, pady=5)
        tk.Label(right_frame, text="STATISTICHE", font=("Helvetica", 12, "bold"), bg="#f0f0f0").pack(pady=(10, 5))
        
        self.lbl_stat_folders = tk.Label(right_frame, text="Cartelle incluse: 0", bg="#f0f0f0", anchor="w")
        self.lbl_stat_folders.pack(fill=tk.X)
        
        self.lbl_stat_files = tk.Label(right_frame, text="File inclusi: 0", bg="#f0f0f0", anchor="w")
        self.lbl_stat_files.pack(fill=tk.X)

        self.lbl_stat_excluded = tk.Label(right_frame, text="Elementi esclusi: 0", bg="#f0f0f0", fg="red", anchor="w")
        self.lbl_stat_excluded.pack(fill=tk.X)

        # --- PANNELLO CENTRALE (Treeview) ---
        tree_frame = tk.Frame(self.root, padx=10, pady=10)
        tree_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        tree_scroll_y = tk.Scrollbar(tree_frame)
        tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        tree_scroll_x = tk.Scrollbar(tree_frame, orient='horizontal')
        tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Colonne aggiunte per i metadati
        cols = ("Tipo", "Stato", "Dimensione", "Data Modifica", "Estensione")
        self.tree = ttk.Treeview(tree_frame, columns=cols, selectmode="browse",
                                 yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
        
        self.tree.heading("#0", text="Nome File / Cartella")
        self.tree.heading("Tipo", text="Tipo")
        self.tree.heading("Stato", text="Stato")
        self.tree.heading("Dimensione", text="Peso")
        self.tree.heading("Data Modifica", text="Ultima Modifica")
        self.tree.heading("Estensione", text="Est.")
        
        self.tree.column("#0", width=350, minwidth=200)
        self.tree.column("Tipo", width=70, anchor=tk.CENTER)
        self.tree.column("Stato", width=70, anchor=tk.CENTER)
        self.tree.column("Dimensione", width=80, anchor=tk.E)
        self.tree.column("Data Modifica", width=120, anchor=tk.CENTER)
        self.tree.column("Estensione", width=60, anchor=tk.CENTER)
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        tree_scroll_y.config(command=self.tree.yview)
        tree_scroll_x.config(command=self.tree.xview)
        
        self.tree.tag_configure('excluded', foreground='gray', font='tkDefaultFont 9 overstrike')
        self.tree.tag_configure('included', foreground='black')
        self.tree.tag_configure('folder', font='tkDefaultFont 9 bold')
        self.tree.tag_configure('symlink', foreground='blue')
        
        self.tree.bind("<Double-1>", self.toggle_item)

    # --- FUNZIONI DI UTILITA' ---
    
    def format_size(self, size_bytes):
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes / 1024:.1f} KB"
        else:
            return f"{size_bytes / (1024 * 1024):.2f} MB"

    def is_hidden_or_system(self, filepath):
        name = os.path.basename(filepath)
        if name.startswith('.') or name.lower() in ['thumbs.db', 'desktop.ini']:
            return True
        if platform.system() == 'Windows':
            try:
                import ctypes
                attrs = ctypes.windll.kernel32.GetFileAttributesW(str(filepath))
                if attrs != -1 and bool(attrs & (2 | 4)):
                    return True
            except:
                pass
        return False

    def update_statistics(self):
        folders, files, excluded = 0, 0, 0
        for item, is_included in self.included_items.items():
            if not is_included:
                excluded += 1
            else:
                item_type = self.tree.item(item, "values")[0]
                if item_type == "Cartella":
                    folders += 1
                elif item_type == "File":
                    files += 1
                    
        self.lbl_stat_folders.config(text=f"Cartelle incluse: {folders}")
        self.lbl_stat_files.config(text=f"File inclusi: {files}")
        self.lbl_stat_excluded.config(text=f"Elementi esclusi: {excluded}")

    # --- NAVIGAZIONE E POPOLAMENTO ---

    def load_directory(self):
        dir_path = filedialog.askdirectory(title="Seleziona la cartella principale")
        if not dir_path:
            return
            
        self.tree.delete(*self.tree.get_children())
        self.included_items.clear()
        
        root_node = self.tree.insert("", "end", text=os.path.basename(dir_path), 
                                     values=("Cartella", "Incluso", "", "", ""), tags=('included', 'folder'))
        self.included_items[root_node] = True
        
        self.populate_tree(root_node, dir_path)
        self.tree.item(root_node, open=True)
        self.update_statistics()

    def populate_tree(self, parent_node, parent_path):
        try:
            items = os.listdir(parent_path)
        except PermissionError:
            return
            
        items.sort(key=lambda x: (not os.path.isdir(os.path.join(parent_path, x)), x.lower()))
        
        for item in items:
            item_path = os.path.join(parent_path, item)
            
            if self.is_hidden_or_system(item_path):
                continue
                
            is_symlink = os.path.islink(item_path)
            is_dir = os.path.isdir(item_path)
            
            item_type = "Cartella" if is_dir else "File"
            if is_symlink:
                item_type = "Link/Alias"
            
            size_str, date_str, ext_str = "", "", ""
            if not is_dir and not is_symlink:
                try:
                    size_str = self.format_size(os.path.getsize(item_path))
                    mtime = os.path.getmtime(item_path)
                    date_str = datetime.datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M")
                    ext_str = os.path.splitext(item)[1].lower()
                except Exception:
                    pass
            
            tags = ['included']
            if is_dir: tags.append('folder')
            if is_symlink: tags.append('symlink')
            
            node = self.tree.insert(parent_node, "end", text=item, 
                                    values=(item_type, "Incluso", size_str, date_str, ext_str), 
                                    tags=tuple(tags))
            self.included_items[node] = True
            
            if is_dir and not is_symlink:
                self.populate_tree(node, item_path)

    # --- AZIONI UTENTE ---

    def toggle_item(self, event):
        item = self.tree.identify_row(event.y)
        if not item:
            return
            
        current_state = self.included_items[item]
        self.cascade_state(item, not current_state)
        self.update_statistics()

    def cascade_state(self, item, state):
        self.included_items[item] = state
        values = list(self.tree.item(item, "values"))
        values[1] = "Incluso" if state else "Escluso"
        
        is_folder = (values[0] == "Cartella")
        state_tag = 'included' if state else 'excluded'
        
        tags = [state_tag]
        if is_folder: tags.append('folder')
        if values[0] == "Link/Alias": tags.append('symlink')
        
        self.tree.item(item, values=values, tags=tuple(tags))
        
        for child in self.tree.get_children(item):
            self.cascade_state(child, state)

    def set_tree_state(self, item, is_open):
        self.tree.item(item, open=is_open)
        for child in self.tree.get_children(item):
            self.set_tree_state(child, is_open)

    def expand_all(self):
        for item in self.tree.get_children():
            self.set_tree_state(item, True)

    def collapse_all(self):
        for item in self.tree.get_children():
            self.set_tree_state(item, False)
            self.tree.item(item, open=True)

    # --- ESPORTAZIONE DATI ---

    def get_max_depth(self, item, current_depth=0):
        if not self.included_items.get(item, False):
            return current_depth
        children = self.tree.get_children(item)
        if not children:
            return current_depth
        return max((self.get_max_depth(child, current_depth + 1) for child in children), default=current_depth)

    def build_export_rows(self, item, current_depth, max_depth, rows):
        if not self.included_items.get(item, False):
            return
            
        text = self.tree.item(item, "text")
        values = self.tree.item(item, "values")
        item_type, _, size, mdate, ext = values
        
        row = [""] * max_depth
        row[current_depth] = text
        row.extend([item_type, size, mdate, ext])
        
        rows.append(row)
        
        for child in self.tree.get_children(item):
            self.build_export_rows(child, current_depth + 1, max_depth, rows)

    def prepare_export_data(self):
        roots = self.tree.get_children()
        if not roots:
            messagebox.showwarning("Attenzione", "Nessun dato da esportare.")
            return None, None
            
        root_node = roots[0]
        max_depth = self.get_max_depth(root_node, 0) + 1
        
        headers = [f"Livello {i}" for i in range(max_depth)] + ["Tipo", "Dimensione", "Data Modifica", "Estensione"]
        rows = []
        self.build_export_rows(root_node, 0, max_depth, rows)
        
        return headers, rows

    def export_csv(self):
        headers, rows = self.prepare_export_data()
        if not headers: return
            
        save_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")], title="Salva CSV")
        if not save_path: return
        
        try:
            with open(save_path, mode='w', newline='', encoding='utf-8-sig') as file:
                writer = csv.writer(file, delimiter=';')
                writer.writerow(headers)
                writer.writerows(rows)
            messagebox.showinfo("Successo", f"CSV salvato in:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            messagebox.showerror("Libreria mancante", "Per esportare in Excel devi installare openpyxl.\nApri il terminale e digita:\npip install openpyxl")
            return

        headers, rows = self.prepare_export_data()
        if not headers: return
            
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], title="Salva Excel")
        if not save_path: return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Foliario"

            # Scrivi Intestazioni
            ws.append(headers)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            # Trova l'indice della colonna "Tipo"
            try:
                tipo_idx = headers.index("Tipo")
            except ValueError:
                tipo_idx = -1

            folder_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            folder_font = Font(bold=True)

            # Scrivi i dati riga per riga e formatta al volo
            for row_idx, row_data in enumerate(rows, start=2): 
                ws.append(row_data)
                
                if tipo_idx != -1 and row_data[tipo_idx] == "Cartella":
                    for col_idx in range(1, len(row_data) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = folder_fill
                        cell.font = folder_font
                        
            # Allarga le colonne A-E
            for col_letter in ['A', 'B', 'C', 'D', 'E']:
                ws.column_dimensions[col_letter].width = 25
                
            wb.save(save_path)
            messagebox.showinfo("Successo", f"Excel generato e formattato con successo in:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Errore", f"Si è verificato un errore:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = FoliarioApp(root)
    root.mainloop()